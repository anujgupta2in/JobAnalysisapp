import pandas as pd
import os
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
import plotly.express as px
import plotly.graph_objects as go

def process_csv_file(file):
    """Process a single CSV file and extract relevant information."""
    try:
        # Extract date from filename using regex
        filename = file.name
        date_match = re.search(r'\b(\d{2})(\d{2})(\d{4})\b', filename)
        formatted_date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}" if date_match else "Unknown"
        # Read CSV file
        df = pd.read_csv(file)
        # Identify the Vessel column
        vessel_column = next((col for col in df.columns if 'vessel' in col.lower()), None)
        vessel_name = df[vessel_column].iloc[0] if vessel_column else "Vessel column not found"
        # Identify the Job Status column
        status_column = next((col for col in df.columns if 'status' in col.lower()), None)
        # Count total jobs
        job_count = len(df)
        # Count jobs with Job Status == 'New'
        if status_column:
            df[status_column] = df[status_column].astype(str).str.strip()
            new_job_count = (df[status_column] == 'New').sum()
        else:
            new_job_count = 0
        return {
            'File Name': filename,
            'Vessel Name': vessel_name,
            'Total Count of Jobs': job_count,
            'New Job Count': new_job_count,
            'Date Extracted from File Name': formatted_date
        }
    except Exception as e:
        return {
            'File Name': filename if 'filename' in locals() else 'Unknown',
            'Vessel Name': 'Error',
            'Total Count of Jobs': 'Error',
            'New Job Count': 'Error',
            'Date Extracted from File Name': formatted_date if 'formatted_date' in locals() else 'Unknown'
        }

def create_vessel_job_distribution_chart(df):
    """Create a bar chart showing job distribution across vessels for individual files."""
    # Sort data by date to maintain chronological order
    df = df.sort_values('Date Extracted from File Name')
    
    fig = go.Figure()
    
    # Add total jobs bars
    fig.add_trace(go.Bar(
        name='Total Jobs',
        x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
        y=df['Total Count of Jobs'],
        marker_color='#1F4E78'
    ))
    
    # Add new jobs bars
    fig.add_trace(go.Bar(
        name='New Jobs',
        x=[f"{row['Vessel Name']} - {row['File Name']}" for _, row in df.iterrows()],
        y=df['New Job Count'],
        marker_color='#F63366'
    ))
    
    # Update layout with improved readability
    fig.update_layout(
        title='Job Distribution by Vessel and File',
        xaxis_title='Vessel - File',
        yaxis_title='Number of Jobs',
        barmode='group',
        height=500,  # Increased height for better visibility
        showlegend=True,
        xaxis=dict(
            tickangle=45,  # Angled labels for better readability
            tickmode='array',
            ticktext=[f"{row['Vessel Name']}<br>{row['File Name']}" for _, row in df.iterrows()],
            tickvals=list(range(len(df)))
        ),
        margin=dict(b=150)  # Increased bottom margin for rotated labels
    )
    
    return fig

def create_jobs_timeline_chart(df):
    """Create a line chart showing job trends over time."""
    timeline_data = df.groupby('Date Extracted from File Name').agg({
        'Total Count of Jobs': 'sum',
        'New Job Count': 'sum'
    }).reset_index()
    
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=timeline_data['Date Extracted from File Name'],
        y=timeline_data['Total Count of Jobs'],
        name='Total Jobs',
        line=dict(color='#1F4E78', width=2)
    ))
    fig.add_trace(go.Scatter(
        x=timeline_data['Date Extracted from File Name'],
        y=timeline_data['New Job Count'],
        name='New Jobs',
        line=dict(color='#F63366', width=2)
    ))
    
    fig.update_layout(
        title='Job Trends Over Time',
        xaxis_title='Date',
        yaxis_title='Number of Jobs',
        height=400,
        showlegend=True
    )
    return fig

def create_jobs_pie_chart(df):
    """Create a pie chart showing the proportion of new vs. existing jobs."""
    total_jobs = df['Total Count of Jobs'].sum()
    new_jobs = df['New Job Count'].sum()
    existing_jobs = total_jobs - new_jobs
    
    fig = go.Figure(data=[go.Pie(
        labels=['New Jobs', 'Existing Jobs'],
        values=[new_jobs, existing_jobs],
        hole=.4,
        marker_colors=['#F63366', '#1F4E78']
    )])
    
    fig.update_layout(
        title='New vs. Existing Jobs Distribution',
        height=400,
        showlegend=True
    )
    return fig

def create_excel_report(df):
    """Create a formatted Excel report from the DataFrame."""
    output = BytesIO()
    
    # Save DataFrame to Excel
    df.to_excel(output, index=False)
    
    # Load workbook for formatting
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Cell borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    
    # Define orange fill for conditional formatting (duplicates)
    orange_fill = PatternFill(start_color="FFB266", end_color="FFB266", fill_type="solid")
    dxf = DifferentialStyle(fill=orange_fill)
    
    # Create rule for duplicate values in Vessel Name column
    dup_rule = Rule(type="duplicateValues", dxf=dxf, stopIfTrue=False)
    ws.conditional_formatting.add(f'B2:B{ws.max_row}', dup_rule)
    
    # Alternating row colors
    gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    for row in range(2, ws.max_row + 1, 2):
        for cell in ws[row]:
            cell.fill = gray_fill
    
    # Add Excel table with formatting
    table_ref = f"A1:E{ws.max_row}"
    table = Table(displayName="JobSummaryTable", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to BytesIO
    output_formatted = BytesIO()
    wb.save(output_formatted)
    output_formatted.seek(0)
    
    return output_formatted
